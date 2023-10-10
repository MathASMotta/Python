from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl

cantor = input('Coloque o nome de um cantor aqui (com a primeira letra maiuscula e acentos): ')

# entrar no site
driver = webdriver.Chrome()
driver.get('https://www.cifraclub.com.br/')
driver.set_window_size(1920,1080)
sleep(10)

# colocar nome do artista na caixa de pesquisa
barra_pesquisa = driver.find_element(By.XPATH, "//input[@id='js-h-search']")
barra_pesquisa.send_keys(cantor)

# clicar em pesquisar
botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='header-searchButton']")
botao_pesquisar.click()
sleep(10)

# clicar no link ao qual está o nome do artista
link_cantor = driver.find_element(By.XPATH, "//a[@class='gs-title']")
link_cantor.click()
sleep(10)

# pegar o nome do artista
nome_cantor = driver.find_element(By.XPATH, "//span[@id='js-artistName']")
nome_cantor = nome_cantor.text

# 10 musicas mais acessadas
musicas = driver.find_elements(By.XPATH, "//ol[@id='js-a-t']//a[contains(@class,'art_music-link')]//div//div//div")
lista_musicas = []
for musica in musicas:
    lista_musicas.append(musica.text)

# colocar as informações em uma planilha excel
workbook = openpyxl.load_workbook('dados.xlsx')
try:
    #pagina existente
    pagina_cantor = workbook[cantor]
    #criar colunas
    pagina_cantor['A1'].value = "Cantor"
    pagina_cantor['B1'].value = "Top 10 Musicas"
    #adicionar nome do cantor
    pagina_cantor['A2'].value = cantor
    #adicionar musicas
    for index, linha in enumerate(pagina_cantor.iter_rows(min_row=2,max_row=len(lista_musicas),min_col=2,max_col=2)):
        for celula in linha:
            celula.value = lista_musicas[index]
    workbook.save('dados.xlsx')
except Exception as error:
    #pagina inexistente
    workbook.create_sheet(cantor)
    pagina_cantor = workbook[cantor]
    #criar colunas
    pagina_cantor['A1'].value = "Cantor"
    pagina_cantor['B1'].value = "Top 10 Musicas"
    #adicionar nome do cantor
    pagina_cantor['A2'].value = cantor
    #adicionar musicas
    for index, linha in enumerate(pagina_cantor.iter_rows(min_row=2,max_row=len(lista_musicas),min_col=2,max_col=2)):
        for celula in linha:
            celula.value = lista_musicas[index]
    workbook.save('dados.xlsx')